from flask import Flask, request, jsonify
import csv
import os
from datetime import datetime, timezone
from bson import ObjectId
from openpyxl import load_workbook
import base64
from io import BytesIO

OPERATIONS_CACHE = {}

app = Flask(__name__)


def utc_mongo_now():
    return int(datetime.now(timezone.utc).timestamp() * 1000)

class DataManagerPython:

    def leer_archivo(self, file_bytes, filename, delimiter=","):
        ext = os.path.splitext(filename)[1].lower()

        if ext in [".csv", ".txt"]:
            for row in csv.reader(file_bytes.decode("utf-8", errors="ignore").splitlines(), delimiter=delimiter):
                yield row

        elif ext == ".xlsx":
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active

            for row in ws.iter_rows(values_only=True):
                yield ["" if v is None else str(v).strip() for v in row]

        else:
            raise Exception(f"Formato no soportado: {ext}")

    def get_trim(self, v):
        return v.strip() if isinstance(v, str) else v

    def fila_tiene_datos(self, fila):
        return any(v not in ("", None) for v in fila)

    def aplicar_filtros(self, data_row, campos_filtro):
        if not campos_filtro:
            return True
        for campo, lookup in campos_filtro.items():
            if campo not in data_row:
                return False
            if data_row[campo] == "" or data_row[campo] not in lookup:
                return False
        return True

    def save_registro(self, data, data_config, id_item, config_recalc):
        data_config_fields = data_config["field"]
        lock_fields = config_recalc["btn_lock"]
        required_fields = config_recalc["btn_asterisk"]

        errores = []
        operations = []
        now = utc_mongo_now()
        for index, registro_data in enumerate(data):
            document = {
                "id_item": id_item,
                "status": "lq",
                "updated_at": now
            }

            result = 4
            for field in required_fields:
                valor = registro_data.get(field)
                if valor is None or (isinstance(valor, str) and valor.strip() == ""):
                    result = 3
                    break

            if result != 4:
                errores.append([index + 1, result])
                continue

            filter_query = {}

            cid = registro_data.get("cid")
            if cid:
                try:
                    filter_query["_id"] = ObjectId(cid)
                except:
                    errores.append([index + 1, 2])
                    continue

            elif lock_fields:
                for f in lock_fields:
                    filter_query[f] = registro_data.get(f, "")
                filter_query["id_item"] = id_item
                filter_query["status"] = "lq"

            for val in data_config_fields:
                campo = val["key_ord"]
                document[campo] = self.get_trim(registro_data.get(campo, ""))

            if filter_query:
                operations.append({
                    "updateOne": [
                        filter_query,
                        {"$set": document,"$setOnInsert": {"created_at": now}},
                        {"upsert": True}
                    ]
                })
            else:
                document["created_at"] = now
                operations.append({"insertOne": [ document ]})

        return {"result": result if not errores else errores, "operations": operations}

    def procesar_lote(self, batch_data, data_config, id_item, array_error, processed, config_recalc):
        result = self.save_registro(batch_data, data_config, id_item, config_recalc)
        errores = result.get("result", [])
        if isinstance(errores, list):
            for err in errores:
                array_error.append([processed + err[0], err[1]])
        return len(batch_data), result.get("operations")

    def process_save_file(self, req):
        data = req["data"]
        data_config = req["dataConfig"]
        config_recalc = req["configPrecalculada"]
        filename = req["filename"]

        file_bytes = base64.b64decode(req["file_base64"], validate=True)

        id_item = data["id"]
        campos = [c["key_ord"] for c in data_config["field"]]

        header_position = int(data.get("header_position", 1))
        tiene_header = str(data.get("header", "")).upper() == "SI"
        delimiter = data.get("delimiter", ",")

        batch_size = 5000
        batch = []
        array_error = []
        processed = 0
        all_operations = []

        campos_filtro = {}
        for c in data_config["field"]:
            filtros = c.get("filter_field", [])
            if filtros:
                campos_filtro[c["key_ord"]] = {v: True for v in filtros}

        rows = list(self.leer_archivo(file_bytes, filename, delimiter))
        info_headers = {}
        headers_row = None

        if tiene_header:
            for i, row in enumerate(rows, start=1):
                if i == header_position:
                    headers_row = [str(h).strip() for h in row]
                    break
            if not headers_row:
                return { "status": "error", "message": "No se encontró la fila de cabecera indicada."}
            config_header = [c["key_ord"] for c in data_config["field"]]
            headers_lookup = {h.lower(): idx for idx, h in enumerate(headers_row)}

            missing = []
            for header in config_header:
                key = header.strip().lower()
                if key not in headers_lookup:
                    missing.append(header)
                else:
                    info_headers[header] = headers_lookup[key]
            if missing:
                return { "status": "error", "message": "Error en el formato del archivo. Faltan columnas obligatorias.", "missing_headers": missing }

        else:
            info_headers = {c["key_ord"]: idx for idx, c in enumerate(data_config["field"])}

        reader = self.leer_archivo(file_bytes, filename, delimiter)
        for i, fila in enumerate(reader, start=1):
            if i < header_position:
                continue
            if tiene_header and i == header_position:
                continue
            if not self.fila_tiene_datos(fila):
                continue

            fila = list(fila)
            if len(fila) < len(campos):
                fila.extend([""] * (len(campos) - len(fila)))

            data_row = {}
            if tiene_header:
                for c in campos:
                    key = c
                    if key not in info_headers:
                        data_row[c] = ""
                    else:
                        pos = info_headers[key]
                        data_row[c] = fila[pos] if pos < len(fila) else ""
            else:
                for idx, c in enumerate(campos):
                    data_row[c] = fila[idx] if idx < len(fila) else ""

            if not self.aplicar_filtros(data_row, campos_filtro):
                continue

            batch.append(data_row)

            if len(batch) >= batch_size:
                cant, ops = self.procesar_lote(
                    batch, data_config, id_item, array_error, processed, config_recalc
                )
                processed += cant
                if ops:
                    all_operations.append(ops)
                batch = []

        if batch:
            cant, ops = self.procesar_lote(
                batch, data_config, id_item, array_error, processed, config_recalc
            )
            processed += cant
            if ops:
                all_operations.append(ops)

        OPERATIONS_CACHE[id_item] = all_operations

        return {
            "status": "ok" if not array_error else "errores",
            "total_processed": processed,
            "errors": array_error,
            "operations_available": True,
            "total_chunks": len(all_operations),
        }


@app.route("/procesar", methods=["POST"])
def procesar():
    try:
        req = request.json
        dm = DataManagerPython()
        result = dm.process_save_file(req)
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 2

@app.get("/operations/<id_item>/page/<int:page>")
def get_operations(id_item, page):
    if id_item not in OPERATIONS_CACHE:
        return jsonify({"status": "error", "message": "ID no encontrado"})
    ops = OPERATIONS_CACHE[id_item]
    if page < 0 or page >= len(ops):
        return jsonify({"status": "error", "message": "Página fuera de rango"})
    return jsonify({
        "page": page,
        "total_pages": len(ops),
        "operations": ops[page]
    })


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8080)
